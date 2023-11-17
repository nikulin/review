import datetime
import re

import openpyxl
import requests
import io
import pyodbc

RUSSIAN_MONTHS_NAMES = {
    'январь': 1,
    'февраль': 2,
    'март': 3,
    'апрель': 4,
    'май': 5,
    'июнь': 6,
    'июль': 7,
    'август': 8,
    'сентябрь': 9,
    'октябрь': 10,
    'ноябрь': 11,
    'декабрь': 12,
}
OUTPUT_DATE_FORMAT = '%d.%m.%Y'

REGIONS_TABLE = 'Mortgage_Regions'
PARAMETERS_TABLE = 'Mortgage_Parameters'
DATA_NAMES_TABLE = 'Mortgage_Data_Names'

#  Коды регионов отсюда
#  https://rosstat.gov.ru/bgd/regl/b09_16/IssWWW.exe/Stg/10-00.htm
REGIONS = (
    (1, 'Российская Федерация', 'The Russian Federation'),
    (2, 'Центральный федеральный округ', 'The Central Federal District'),
    (3, 'Белгородская область', 'Belgorod region'),
    (4, 'Брянская область', 'Bryansk region '),
    (5, 'Владимирская область', 'Vladimir region'),
    (6, 'Воронежская область', 'Voronezh region'),
    (7, 'Ивановская область', 'Ivanovo region'),
    (8, 'Калужская область', 'Kaluga region'),
    (9, 'Костромская область', 'Kostroma region'),
    (10, 'Курская область', 'Kursk region'),
    (11, 'Липецкая область', 'Lipetzk region'),
    (12, 'Московская область', 'Moscow region'),
    (13, 'Орловская область', 'Oryol region'),
    (14, 'Рязанская область', 'Ryazan region '),
    (15, 'Смоленская область', 'Smolensk region'),
    (16, 'Тамбовская область', 'Tambov region'),
    (17, 'Тверская область', 'Tver region'),
    (18, 'Тульская область', 'Tula region'),
    (19, 'Ярославская область', 'Yaroslavl region'),
    (20, 'г. Москва', 'The City of Moscow'),
    (21, 'Северо-Западный федеральный округ', 'The North West Federal District'),
    (22, 'Республика Карелия', 'Republic of Karelia '),
    (23, 'Республика Коми', 'Republic of Komi'),
    (24, 'Архангельская область', 'Arkhangelsk region'),
    (25, 'в том числе Ненецкий автономный округ', 'Nenets autonomous district'),
    (
        240,
        'Архангельская область без данных по Ненецкому автономному округу',
        'Arkhangelsk region w/o Nenets autonomous district',
    ),
    (26, 'Вологодская область', 'Vologda region'),
    (27, 'Калининградская область', 'Kaliningrad region'),
    (28, 'Ленинградская область', 'Leningrad region'),
    (29, 'Мурманская область', 'Murmansk region'),
    (30, 'Новгородская область', 'Novgorod region '),
    (31, 'Псковская область', 'Pskov region'),
    (32, 'г. Санкт-Петербург', 'The City of Sankt-Petersburg'),
    (33, 'Южный федеральный округ', 'The South Federal District'),
    (34, 'Республика Адыгея', 'Republic of Adygeya'),
    (35, 'Республика Дагестан', 'Republic of Dagestan'),
    (36, 'Республика Ингушетия', 'Republic of Ingushetia'),
    (37, 'Кабардино-Балкарская Республика Kabardino-Balkarian', 'Republic'),
    (38, 'Республика Калмыкия', 'Republic of Kalmykia '),
    (39, 'Карачаево-Черкесская Республика', 'Karachaevo-Chercessian Republic'),
    (40, 'Республика Северная Осетия - Алания', 'Republic of North Ossetia - Alania'),
    (41, 'Чеченская Республика', 'Chechen Republic'),
    (42, 'Краснодарский край', 'Krasnodar territory'),

STORE_PARAMETERS = 1
STORE_REGIONS = 2

BASE_URL = 'https://www.cbr.ru/vfs/statistics/BankSector/Mortgage/'
FILES = {
    '02_02_Mortgage.xlsx': STORE_PARAMETERS,
    '02_03_Scpa_mortgage.xlsx': STORE_PARAMETERS,
    '02_10_Quantity_mortgage.xlsx': STORE_REGIONS,
    # '02_15_Quantity_scpa_mortgage.xlsx': STORE_REGIONS,
    # '02_11_New_loans_mortgage.xlsx': STORE_REGIONS,
    # '02_16_New_loans_scpa_mortgage.xlsx': STORE_REGIONS,
    # '02_14_Debt_mortgage.xlsx': STORE_REGIONS,
    # '02_18_Debt_scpa_mortgage.xlsx': STORE_REGIONS,
    # '02_13_Rates_mortgage.xlsx': STORE_REGIONS,
    # '02_17_Rates_scpa_mortgage.xlsx': STORE_REGIONS,
}


def is_table_exists(cursor, table_name):
    queries = {
        'Microsoft SQL Server': "SELECT * FROM information_schema.tables WHERE table_name = '{table_name}'",
    }

    server_type = cursor.connection.getinfo(pyodbc.SQL_DBMS_NAME)
    cursor.execute(queries[server_type].format(table_name=table_name))
    return cursor.fetchone() is not None


def make_table_name(string: str, prefix: str = '') -> str:
    name = re.sub('_(.)', lambda match: match.group(0), string[3:string.rfind('.')])
    if prefix:
        name = f'{prefix}_{name}'
    if name[0].isdigit():
        name = f'T{name}'
    return name.lower()


def clean_string(string) -> str:
    """
    Remove doubled spaces and capitalize the string
    :param string: input string
    :return:cleaned string
    """
    return ' '.join(string.strip().split()).capitalize()


def convert_date(string: str) -> str:
    """
    Convert russian dates like "Август 2023" to 01.08.2023
    :param string:
    :return:
    """
    try:
        string = string.lower()
        month_name, year = string.split()
        prefix = "01.{month:02d}".format(month=RUSSIAN_MONTHS_NAMES[month_name])
        string = string.replace(month_name, prefix).replace(' ', '.')
    except ValueError:
        pass

    return string


if __name__ == '__main__':
    connection = pyodbc.connect(
        "DRIVER={{ODBC Driver 17 for SQL Server}};"
        "SERVER={server};"
        "DATABASE={db};"
        "UID={user};"
        "PWD={password};"
        "PORT=1433;"
        "TRUSTED_CONNECTION=no".format(
            server='localhost',
            db='CBRF',
            user='sa',
            password='SuperSTRONG-password=1433',
        )
    )
    cursor = connection.cursor()

    table_count = sum((
        is_table_exists(cursor, REGIONS_TABLE),
        is_table_exists(cursor, PARAMETERS_TABLE),
        is_table_exists(cursor, DATA_NAMES_TABLE),
    ))

    if 0 < table_count < 3:
        raise ValueError('Some tables dont exist, check they manually')

    this_is_first_run = table_count == 0

    if this_is_first_run:
        cursor.execute("""
            CREATE TABLE {table_name}
            (
                id INTEGER PRIMARY KEY,
                title VARCHAR(255) NOT NULL UNIQUE,
                title_eng VARCHAR(255) NOT NULL UNIQUE,
            )
        """.format(table_name=REGIONS_TABLE.lower()))

        cursor.executemany(
            f"INSERT INTO {REGIONS_TABLE} VALUES (?, ?, ?)",
            REGIONS,
        )
        connection.commit()

        for table_name in (PARAMETERS_TABLE, DATA_NAMES_TABLE):
            cursor.execute("""
                CREATE TABLE {table_name}
                (
                    id INTEGER IDENTITY(1, 1) PRIMARY KEY,
                    -- code INTEGER NOT NULL UNIQUE CHECK (code > 0),
                    name VARCHAR(255) NOT NULL UNIQUE
                )
            """.format(table_name=table_name.lower()))

    for file, store_type in FILES.items():
        if this_is_first_run:
            cursor.execute(
                """
                CREATE TABLE {table_name}
                (
                    id INTEGER IDENTITY(1, 1) PRIMARY KEY,
                    data_name_id INTEGER NOT NULL CHECK (data_name_id > 0),
                    {link_name} INTEGER NOT NULL CHECK ( {link_name} > 0 ),
                    value DECIMAL(18, 3) NOT NULL,
                    date DATE NOT NULL,
                    created_at DATETIME2 NOT NULL,
                    INDEX IDX_date NONCLUSTERED(date),
                    INDEX IDX_data_per_date NONCLUSTERED(data_name_id, {link_name}, date),
                    CONSTRAINT UC_data_date_value_{table_name} UNIQUE(data_name_id, {link_name}, date, value)
                )
                """.format(
                    table_name=make_table_name(file),
                    link_name='region_id' if store_type == STORE_REGIONS else 'parameter_id',
                )
            )

    cursor.execute(f"SELECT id, name FROM {DATA_NAMES_TABLE}")
    data_names = {name.lower(): code for code, name in cursor.fetchall()}
    cursor.execute(f"SELECT id, name FROM {PARAMETERS_TABLE}")
    parameters = {name.lower(): code for code, name in cursor.fetchall()}
    cursor.execute(f"SELECT id, title FROM {REGIONS_TABLE}")
    regions = {name.lower(): code for code, name in cursor.fetchall()}

    for file, store_type in FILES.items():
        response = requests.get(f'{BASE_URL}{file}')
        response.raise_for_status()
        
        with io.BytesIO(response.content) as xls_file:
            wb = openpyxl.load_workbook(filename=xls_file)

            link_table = REGIONS_TABLE if store_type == STORE_REGIONS else PARAMETERS_TABLE

            for i, sheet in enumerate(wb.worksheets):
                print('SHEET:', sheet.title)
                data_name = clean_string(sheet.title)
                if data_name not in data_names:
                    try:
                        cursor.execute(f"INSERT INTO {DATA_NAMES_TABLE} (name) VALUES (?)", data_name)
                        connection.commit()
                    except pyodbc.IntegrityError:
                        pass

                    cursor.execute(f"SELECT id FROM {DATA_NAMES_TABLE} WHERE name=?", data_name)
                    last_id, *_ = cursor.fetchone()
                    data_names[data_name.lower()] = last_id

                if store_type == STORE_PARAMETERS:
                    for cell, *_ in sheet[f'A3:A{sheet.max_row}']:
                        parameter_name = clean_string(cell.value)
                        if parameter_name not in parameters:
                            try:
                                cursor.execute(f"INSERT INTO {link_table} (name) VALUES (?)", parameter_name)
                                connection.commit()
                            except pyodbc.IntegrityError:
                                pass

                            cursor.execute(f"SELECT id FROM {link_table} WHERE name=?", parameter_name)
                            last_id, *_ = cursor.fetchone()
                            parameters[parameter_name.lower()] = last_id

                _left_top, right_bottom = sheet.dimensions.split(':')
                for row, (first_column, *_) in zip(sheet[f'B3:{right_bottom}'], sheet[f'A3:A{sheet.max_row}']):
                    link_name = clean_string(first_column.value) if store_type == STORE_PARAMETERS else first_column.value.strip()
                    data = []
                    last_column = openpyxl.utils.get_column_letter(sheet.max_column)
                    for cell, date in zip(row, sheet[f'B2:{last_column}2'][0]):
                        data.append((
                            data_names[data_name.lower()],
                            (parameters if store_type == STORE_PARAMETERS else regions)[link_name.lower()],
                            cell.value if cell.value not in ('0,0', '0,00') else 0,
                            convert_date(date.value),
                        ))
                    try:
                        cursor.executemany("""
                            INSERT INTO {table_name}
                            (data_name_id, {link_name}, value, date, created_at)
                            VALUES (?, ?, ?, ?, SYSDATETIME()) 
                        """.format(
                            table_name=make_table_name(file),
                            link_name='region_id' if store_type == STORE_REGIONS else 'parameter_id',
                        ), data)
                        connection.commit()
                    except (pyodbc.DataError, pyodbc.IntegrityError) as exc:
                        raise ValueError from exc

            connection.commit()
